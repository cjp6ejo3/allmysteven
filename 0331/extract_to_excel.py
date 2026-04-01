# -*- coding: utf-8 -*-
"""
從 github 資料夾內所有 Yahoo 序號查詢結果中，
提取「📱 發送到 Telegram 的獎品 📱」區塊的獎項與網址，
彙整成一個 Excel 檔案。

txp.rs 是否重新連網檢查「已兌換」：見檔案置頂 FORCE_REFRESH_VOUCHER_STATUS；
亦可執行時加參數 --refresh。
"""

import os
import re
import sys
import glob
import time
import unicodedata
from pathlib import Path
from datetime import datetime
from urllib.request import urlopen, Request
from urllib.error import URLError, HTTPError
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
import pandas as pd

# 腳本所在目錄 = github 資料夾
BASE_DIR = Path(__file__).resolve().parent
TXT_PATTERN = str(BASE_DIR / "Yahoo序號連結查詢結果_*.txt")
# 產出檔名包含時間
FILENAME_BASE = "Yahoo獎項網址彙整"
MAIN_EXCEL = BASE_DIR / f"{FILENAME_BASE}.xlsx"
TIMESTAMP = datetime.now().strftime('%Y%m%d%H%M')
OUTPUT_EXCEL = BASE_DIR / f"{FILENAME_BASE}_{TIMESTAMP}.xlsx"
EXPIRY_CACHE = BASE_DIR / "expiry_cache.txt"   # 兌換期間至／使用狀態快取（url -> 資料）
MAX_THREADS = 30  # 多執行緒檢查數量

# ── txp.rs 連網檢查：兌換期間至 + 是否已兌換 ──
# False（預設）：使用 expiry_cache.txt 快取。已快取過的網址不重複連網（省時間）；
#               若曾寫入「未標示已兌換」，之後也不會自動更新，除非改 True 或刪快取列。
# True：略過快取，每筆 txp.rs 都重新下載 HTML 再判斷「已兌換」與到期日（較慢、狀態較新）。
# 命令列加 --refresh 時，本次執行強制等同 True（可蓋過上面常數，不必改檔）。False 是快取
FORCE_REFRESH_VOUCHER_STATUS = False

# ── 獎項標題 → 價錢對照（「統計數量」分頁用）
# 比對時會做 NFKC 正規化並去掉空白／零寬字元。
# 優先「正規化後完全相等」；若無，則「字典 key 與實際標題任一方包含另一方」即視為同品項
# （例如實際標題較長、字典只寫簡稱）。多筆同時符合時取「key 字數最長」者，較不易誤配。
# 未列出的品項價錢留白。
PRICE_LOOKUP = {
    "摩斯漢堡蛋堡套餐（早餐限定）": 30,
    "85度C58元切片蛋糕": 30,
    "85°C58元切片蛋糕": 30,
    "摩斯漢堡超級大麥海洋珍珠堡套餐": 70,
    "肯德基咔啦脆雞脆薯套餐": 40,
    "必勝客夏威夷6吋個人比薩": 40,
    "台酒花雕雞/花雕酸菜牛肉": 30,
    "CoCo都可珍珠鮮奶茶": 30,
    "麥當勞OREO冰炫風": 25,
    "麥當勞勁辣鷄腿堡套餐": 52,
    # 簡體「劲」或來源異寫時仍可對到
    "麥當勞劲辣雞腿堡套餐": 52,
    "Mister Donut經典午茶組": 35,
}


def _norm_title_key(title):
    """與 PRICE_LOOKUP 比對用的正規化標題（去空白、NFKC）。"""
    if title is None or (isinstance(title, float) and pd.isna(title)):
        return ""
    t = unicodedata.normalize("NFKC", str(title)).strip()
    # 零寬字元、各種空白、全形空格
    t = re.sub(r"[\s\u200b\u200c\u200d\ufeff\u3000]+", "", t)
    # 85°C / ℃ 與「85度C」資料來源不一時，統一成「度C」再比對
    t = t.replace("℃", "度C").replace("°C", "度C")
    if not t or t.lower() == "nan":
        return ""
    return t


# 正規化後標題 -> 價錢（多個原始 key 若正規化相同，後者覆蓋前者）
PRICE_BY_NORM = {_norm_title_key(k): v for k, v in PRICE_LOOKUP.items()}


def price_for_title(title):
    """依「標題」查價錢；對不到則回傳空白字串（Excel 顯示為空）。"""
    t = _norm_title_key(title)
    if not t:
        return ""
    # 1) 正規化後完全相等
    v = PRICE_BY_NORM.get(t)
    if v is not None:
        return v
    # 2) 包含比對：字典 key 出現在標題裡，或標題出現在 key 裡（擇最長 key 以免誤配）
    best_key = None
    best_price = None
    for nk, price in PRICE_BY_NORM.items():
        if not nk:
            continue
        if nk in t or t in nk:
            if best_key is None or len(nk) > len(best_key):
                best_key = nk
                best_price = price
    return best_price if best_price is not None else ""


def fetch_voucher_info(url):
    """從 txp.rs 兌換券頁面爬取：兌換期間至、是否已使用。"""
    if "txp.rs" not in url:
        return None, ""
    try:
        req = Request(url, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"})
        with urlopen(req, timeout=15) as resp:
            html = resp.read().decode("utf-8", errors="ignore")
        
        expiry = None
        # 爬取日期格式: 2026.05.23 或 2026/05/23
        m = re.search(r"([\d]{4}[\./][\d]{2}[\./][\d]{2})", html)
        if m:
            expiry = m.group(1).strip().replace("/", ".")
        
        # 檢測「已兌換/已使用」狀態的精準標記 (排除注意事項中的文字)
        used_markers = [
            "usededn",           # Edenred CSS marker
            "StatusOverlayBg",   # Edenred Status overlay
            "stamp_used.png",    # 'Used' stamp image
            "StampStatus",       # Stamp status element
            'class="stamp used"', # Specific CSS class
            'id="balance">剩餘：0' # Balance is zero (specifically in the balance element)
        ]
        is_used = any(x in html for x in used_markers)
        
        used = "已兌換" if is_used else ""
        return expiry, used
    except (URLError, HTTPError, Exception):
        pass
    return None, ""


def load_expiry_cache():
    """載入兌換期間至快取（格式：url\texpiry 或 url\texpiry\tstatus）。"""
    cache = {}  # url -> (expiry, status)
    if EXPIRY_CACHE.exists():
        try:
            for line in EXPIRY_CACHE.read_text(encoding="utf-8").strip().split("\n"):
                if "\t" in line:
                    parts = line.split("\t", 2)
                    url = parts[0].strip()
                    expiry = parts[1].strip() if len(parts) > 1 else ""
                    status = parts[2].strip() if len(parts) > 2 else ""
                    cache[url] = (expiry, status)
        except Exception:
            pass
    return cache


def save_expiry_cache(cache):
    """儲存兌換期間至快取。"""
    lines = []
    for url in sorted(cache.keys()):
        expiry, status = cache[url]
        if status:
            lines.append(f"{url}\t{expiry}\t{status}")
        else:
            lines.append(f"{url}\t{expiry}")
    EXPIRY_CACHE.write_text("\n".join(lines), encoding="utf-8")


def enrich_prizes_with_expiry(prizes_list, verbose=True, force_refresh=False):
    """為每個獎品補充兌換期間至、是否已使用，使用多執行緒加速。"""
    cache = load_expiry_cache() if not force_refresh else {}
    
    # 找出需要檢查的品項
    to_check = []
    for p in prizes_list:
        url = p["link"].strip()
        if "txp.rs" not in url:
            continue
        if url in cache and not force_refresh:
            expiry, status = cache[url]
            p["expiry"] = expiry
            p["used"] = status
        else:
            to_check.append(p)

    if not to_check:
        return prizes_list

    print(f"  -> 共有 {len(to_check)} 筆需要連網檢查狀態 (使用 {MAX_THREADS} 執行緒)...")
    
    updated = False
    
    def process_p(p):
        url = p["link"].strip()
        expiry, used = fetch_voucher_info(url)
        p["expiry"] = expiry if expiry else ""
        p["used"] = used
        if verbose:
            tag = f"[{used}]" if used else "[可兌換]"
            print(f"  取得: {p['title'][:15]}... -> {tag}")
        return url, (p["expiry"], p["used"])

    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        results = list(executor.map(process_p, to_check))
    
    for url, result_val in results:
        cache[url] = result_val
        updated = True

    if updated:
        save_expiry_cache(cache)
    return prizes_list


def parse_telegram_section(content):
    """從單一檔案內容中解析「發送到 Telegram 的獎品」區塊。"""
    marker = "📱 發送到 Telegram 的獎品 📱"
    if marker not in content:
        return None, []

    start = content.find(marker)
    block = content[start:]
    end_marker = "--- 批次流程結束 ---"
    if end_marker in block:
        block = block[: block.find(end_marker)]

    date_m = re.search(r"發送日期:\s*(.+?)(?:\n|$)", block)
    send_date = date_m.group(1).strip() if date_m else ""

    # 獎品項目： [N] Profile X \n 標題: ... \n 時間: ... \n 連結: ...
    pattern = re.compile(
        r"\[\s*(\d+)\s*\]\s*Profile\s*(\d+)\s*\n\s*標題:\s*(.+?)\s*\n\s*時間:\s*(.+?)\s*\n\s*連結:\s*(.+?)(?=\n\s*\[\s*\d|\n\n|\n===|\Z)",
        re.DOTALL,
    )
    prizes = []
    for m in pattern.finditer(block):
        num, profile, title, time_str, link = m.groups()
        prizes.append({
            "num": num.strip(),
            "profile": profile.strip(),
            "title": title.strip(),
            "time": time_str.strip(),
            "link": link.strip(),
        })
    return send_date, prizes


def collect_all_prizes():
    """掃描 github 資料夾內所有 txt，收集 Telegram 區塊的獎項。"""
    files = sorted(glob.glob(TXT_PATTERN))
    all_entries = []

    for fpath in files:
        fname = os.path.basename(fpath)
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                content = f.read()
        except Exception as e:
            print(f"讀取失敗 {fpath}: {e}")
            continue
        send_date, prizes = parse_telegram_section(content)
        if send_date is None:
            continue
        all_entries.append({"file": fname, "send_date": send_date, "prizes": prizes})

    return all_entries


def _sort_and_group_prizes_flat(flat):
    """同類型放一起，依到期日由近到遠排序（移植自 extract_and_upload.py）。"""
    groups = defaultdict(list)
    for p in flat:
        groups[p["標題"]].append(p)
    
    FAR = "9999.99.99"
    # 每組內依到期日排序
    for title in groups:
        groups[title].sort(key=lambda p: p.get("兌換期間至") or FAR)
    
    # 各組依「該組最早到期日」排序
    def group_min_expiry(items):
        expiries = [p.get("兌換期間至") for p in items if p.get("兌換期間至")]
        return min(expiries) if expiries else FAR
    
    sorted_pairs = sorted(groups.items(), key=lambda x: group_min_expiry(x[1]))
    result = []
    for title, items in sorted_pairs:
        result.extend(items)
    return result


def main():
    print("正在掃描 github 資料夾內的 Yahoo 序號查詢結果...")
    entries = collect_all_prizes()
    if not entries:
        print("未找到任何「📱 發送到 Telegram 的獎品 📱」區塊。")
        return

    # 計算總共找到多少項目
    total_count = sum(len(e["prizes"]) for e in entries)
    print(f"共找到 {len(entries)} 個日期的 Telegram 獎項區塊，累計 {total_count} 筆獎項。")
    
    force_refresh = FORCE_REFRESH_VOUCHER_STATUS or ("--refresh" in sys.argv)
    if force_refresh:
        reasons = []
        if FORCE_REFRESH_VOUCHER_STATUS:
            reasons.append("FORCE_REFRESH_VOUCHER_STATUS=True（置頂）")
        if "--refresh" in sys.argv:
            reasons.append("命令列 --refresh")
        print(f"正在重新爬取所有兌換券（含是否已兌換）〔{'；'.join(reasons)}〕...")
    else:
        print(
            "正在爬取兌換期間至與使用狀態（使用 expiry_cache.txt 快取；"
            "置頂改 FORCE_REFRESH_VOUCHER_STATUS=True 或加 --refresh 可強制重抓）..."
        )
    
    # 攤平所有獎項以便檢查與處理
    prizes_list = []
    for rec in entries:
        send_date = rec["send_date"]
        for p in rec["prizes"]:
            p["send_date"] = send_date # 注入發送日期
            prizes_list.append(p)

    # 為每個獎品補充兌換期間至、是否已使用 (多執行緒版)
    prizes_list = enrich_prizes_with_expiry(prizes_list, verbose=True, force_refresh=force_refresh)

    # 建立最終資料集（去重）
    flat_data = []
    seen_urls = set()
    for p in prizes_list:
        url = p["link"].strip()
        if url in seen_urls:
            continue
        seen_urls.add(url)
        
        flat_data.append({
            "標題": p["title"],
            "兌換期間至": p.get("expiry") or "",
            "發送日期": p["send_date"],
            "時間": p["time"],
            "Profile": f"Profile {p['profile']}",
            "連結": p["link"],
            "使用狀態": p.get("used") or ""
        })

    if not flat_data:
        print("沒有可彙整的資料。")
        return

    # 按照原 HTML 的邏輯排序
    flat_data = _sort_and_group_prizes_flat(flat_data)

    # 讀取舊的 Excel 各個分頁中的備註，以防遺漏
    existing_remarks = {} # link -> remark
    if MAIN_EXCEL.exists():
        try:
            print(f"正在從主檔案讀取備註: {MAIN_EXCEL.name}")
            # 讀取所有分頁 (Sheet 字典)
            all_sheets = pd.read_excel(MAIN_EXCEL, sheet_name=None)
            
            # 先處理總彙整 (如果有)，再處理分頁，這樣分頁內容會覆蓋總彙整
            sheet_names = list(all_sheets.keys())
            if '總彙整' in sheet_names:
                sheet_names.remove('總彙整')
                sheet_names.insert(0, '總彙整')
            
            for sheet_name in sheet_names:
                old_df = all_sheets[sheet_name]
                if '連結' in old_df.columns and '備註' in old_df.columns:
                    for _, row in old_df.iterrows():
                        link = str(row['連結']).strip()
                        remark = str(row['備註']).strip()
                        if link and remark and remark != "nan" and remark != "":
                            existing_remarks[link] = remark
        except Exception as e:
            print(f"提示: 無法讀取舊 Excel 備註: {e}")

    # 建立 DataFrame
    df = pd.DataFrame(flat_data)
    
    # 新的欄位順序：#, 備註, 標題, 兌換期間至, 連結, 使用狀態, 發送日期, 時間, Profile
    # 從舊資料對應「備註」
    df["備註"] = df["連結"].apply(lambda x: existing_remarks.get(str(x).strip(), ""))
    
    # 如果「使用狀態」為「已兌換」且原本沒寫備註，才標記成「已兌換」
    df.loc[(df["使用狀態"] == "已兌換") & (df["備註"] == ""), "備註"] = "已兌換"
    
    cols = ["備註", "標題", "兌換期間至", "連結", "使用狀態", "發送日期", "時間", "Profile"]
    df = df[cols]
    # 「總彙整」排序規則：
    # 1) 備註為空白放最上面
    # 2) 備註非空白集中排到最下面
    # 3) 各自維持原本先後順序（穩定分組，不做內容重排）
    _remark_clean = df["備註"].fillna("").astype(str).str.strip()
    _blank_mask = _remark_clean.eq("") | _remark_clean.eq("nan")
    df = pd.concat([df[_blank_mask], df[~_blank_mask]], ignore_index=True)

    df.insert(0, "#", range(1, len(df) + 1))

    # 建立統計數量的 DataFrame
    # 統計規則：模仿 Excel COUNTA 邏輯。只要「備註」有內容（長度 > 0）或已兌換，就不計數。
    print("正在執行精確過濾（扣除已備註項目）...")
    
    def get_countable_value(row):
        # 取得備註，清洗掉 NaN、前後空白與隱形字元
        rem = str(row.get("備註", "")).strip().replace("nan", "")
        # 取得使用狀態
        stat = str(row.get("使用狀態", "")).strip().replace("nan", "")
        
        # 如果「備註」有任何內容（COUNTA 為 1），這筆算 0 (不計入數量)
        if len(rem) > 0:
            return 0
        # 如果狀態已經標記為「已兌換」，算 0 (不計入數量)
        if stat == "已兌換":
            return 0
        # 只有真正全空的才算 1
        return 1

    # 建立一個臨時的計數欄位
    df["_tmp_count"] = df.apply(get_countable_value, axis=1)
    
    # 依照標題與到期日分組，並將剛才生成的計數欄位加總
    stats_df = df.groupby(["標題", "兌換期間至"], sort=False)["_tmp_count"].sum().reset_index(name="數量")
    stats_df["價錢"] = stats_df["標題"].map(price_for_title)
    # 「總計」欄請在 Excel 手填倍數（例如 1、2）；金額合計 = Σ(價錢 × 總計)，由表尾公式計算。
    stats_df["總計"] = ""
    # 欄位順序：標題、兌換期間至、數量 → 價錢 → 總計
    stats_df = stats_df[["標題", "兌換期間至", "數量", "價錢", "總計"]]
    
    # 移除輔助欄位
    df.drop(columns=["_tmp_count"], inplace=True)
    
    # 匯出到 Excel 的函式
    def save_to_path(path):
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            # 1. 導出一個「總彙整」分頁 (全清單)
            df.to_excel(writer, index=False, sheet_name='總彙整')
            _format_sheet(writer.sheets['總彙整'], df)

            # 2. 導出一個「統計數量」分頁 (統計清單)
            stats_df.to_excel(writer, index=False, sheet_name='統計數量')
            stats_ws = writer.sheets['統計數量']
            _format_sheet(stats_ws, stats_df)
            _add_stats_grand_total_row(stats_ws, n_data_rows=len(stats_df))

            # 3. 依照「標題」拆分分頁
            for title, group in df.groupby("標題", sort=False):
                sheet_name = re.sub(r'[\\/*?:\[\]]', '', title)[:30]
                if not sheet_name: sheet_name = "未命名"
                group = group.copy()
                group["#"] = range(1, len(group) + 1)
                group.to_excel(writer, index=False, sheet_name=sheet_name)
                _format_sheet(writer.sheets[sheet_name], group)

    try:
        # 同時儲存到主檔案與備份檔
        save_to_path(MAIN_EXCEL)
        print(f"✅ 主檔案已更新: {MAIN_EXCEL.name}")
        save_to_path(OUTPUT_EXCEL)
        print(f"✅ 本次備份已存放: {OUTPUT_EXCEL.name}")
    except Exception as e:
        print(f"❌ 寫入 Excel 失敗: {e}")

    print(f"完成。最終彙整出 {len(df)} 筆不重複獎項。")


def _add_stats_grand_total_row(worksheet, n_data_rows):
    """
    「統計數量」表尾：E 欄寫入 SUMPRODUCT(價錢欄 D, 手填倍數欄 E)，即 Σ(價錢×總計)。
    n_data_rows：資料列筆數（不含第 1 列表頭）。
    """
    from openpyxl.styles import Border, Side, Alignment

    if n_data_rows <= 0:
        return

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    last_data_row = 1 + n_data_rows
    sum_row = last_data_row + 1

    c_a = worksheet.cell(row=sum_row, column=1, value="金額合計")
    c_a.border = border
    c_a.alignment = Alignment(vertical="center")

    formula = f"=SUMPRODUCT(D2:D{last_data_row},E2:E{last_data_row})"
    c_e = worksheet.cell(row=sum_row, column=5, value=formula)
    c_e.border = border
    c_e.alignment = Alignment(vertical="center")
    c_e.number_format = "0"


def _format_sheet(worksheet, df):
    """設定工作表的格式，包括自動調整欄寬與框線。"""
    from openpyxl.styles import Border, Side, Alignment
    
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # 自動調寬 (考慮中文字符)
    def calculate_width(s):
        width = 0
        for char in str(s):
            if '\u4e00' <= char <= '\u9fff':  # 中文字符
                width += 2
            else:
                width += 1.1
        return width

    for idx, col in enumerate(df.columns):
        max_len = max(
            df[col].apply(calculate_width).max() if not df[col].empty else 0,
            calculate_width(col)
        ) + 4
        col_letter = chr(65 + idx)
        worksheet.column_dimensions[col_letter].width = min(max_len, 80)
        
        # 套用框線與對齊
        for row in range(1, len(df) + 2):
            cell = worksheet.cell(row=row, column=idx + 1)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            # 如果是「備註」欄位，設定為純文字格式
            if col == "備註":
                cell.number_format = '@'

    # 凍結首列
    worksheet.freeze_panes = 'A2'


if __name__ == "__main__":
    main()

input()
